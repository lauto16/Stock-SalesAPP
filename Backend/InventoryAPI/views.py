from .serializers import (
    ProductSerializer,
    OfferSerializer,
    ProductPagination,
    OfferPagination,
)
from rest_framework.authentication import SessionAuthentication, TokenAuthentication
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from rest_framework.response import Response
from rest_framework.decorators import action
from ProvidersAPI.models import Provider
from rest_framework.views import APIView
from rest_framework import permissions
from SalesAPI.models import SaleItem
from django.http import FileResponse
from rest_framework import viewsets
from django.http import HttpRequest
from .models import Product, Offer
from rest_framework import status
from django.utils import timezone
from openpyxl import Workbook
import os


class ProductValidator:
    """
    Provides validation logic for Product PATCH request data.

    This helper class validates individual fields sent in a partial update
    request and returns a consistent response format indicating whether the
    data is valid. It performs checks for:
      - Empty or invalid text fields
      - Numeric fields (stock, buy price, sell price)
      - Provider validation, including ID format and existence

    All validation errors follow the same structure:
        {"success": False, "error": "<message>"}

    If no issues are found:
        {"success": True, "error": None}
    """

    @staticmethod
    def validate_data(request: HttpRequest) -> dict:
        """
        Validates incoming PATCH data for updating a Product instance.

        The method performs type checks, value constraints, and related-object
        existence checks. It safely handles unexpected errors and wraps them
        in a friendly error message.

        Args:
            request (HttpRequest): The incoming HTTP request containing PATCH data.

        Returns:
            dict: A result dictionary with:
                - success (bool): Whether validation passed.
                - error (str | None): A human-readable error message, or None if valid.
        """
        data = request.data

        try:
            code = data.get("code")
            name = data.get("name")
            stock = data.get("stock")
            buy_price = data.get("buy_price")
            sell_price = data.get("sell_price")
            provider = data.get("provider")
            category = data.get("category")

            if code is not None and not code.strip():
                return {
                    "success": False,
                    "error": "El campo 'código' no puede estar vacío.",
                }

            if name is not None and not name.strip():
                return {
                    "success": False,
                    "error": "El campo 'nombre' no puede estar vacío.",
                }

            if stock is not None:
                try:
                    stock = int(stock)
                except ValueError:
                    return {
                        "success": False,
                        "error": "Stock debe ser un número entero.",
                    }
                if stock < 0:
                    return {"success": False, "error": "Stock no puede ser negativo."}

            if buy_price is not None:
                try:
                    buy_price = float(buy_price)
                except ValueError:
                    return {
                        "success": False,
                        "error": "Precio de compra debe ser un número.",
                    }
                if buy_price <= 0:
                    return {
                        "success": False,
                        "error": "El precio de compra debe ser mayor a 0.",
                    }

            if sell_price is not None:
                try:
                    sell_price = float(sell_price)
                except ValueError:
                    return {
                        "success": False,
                        "error": "Precio de venta debe ser un número.",
                    }
                if sell_price <= 0:
                    return {
                        "success": False,
                        "error": "El precio de venta debe ser mayor a 0.",
                    }

            if provider:
                try:
                    provider_id = int(provider)
                except ValueError:
                    return {
                        "success": False,
                        "error": "El ID del proveedor debe ser un número.",
                    }

                if not Provider.objects.filter(id=provider_id).exists():
                    return {
                        "success": False,
                        "error": "El proveedor especificado no existe.",
                    }

            if category:
                if not category.objects.filter(name=category).exists():
                    return {
                        "success": False,
                        "error": "La categoria especificada no existe.",
                    }

            return {"success": True, "error": None}

        except Exception as e:
            return {"success": False, "error": f"Ocurrió un error inesperado: {str(e)}"}


class ProductViewSet(viewsets.ModelViewSet):
    """
    ViewSet providing CRUD operations and custom actions for Product objects.

    This ViewSet includes:
      - Standard ModelViewSet functionality (list, retrieve, create, update, delete)
      - Custom pagination using ProductPagination
      - Session and token authentication
      - Permission enforcement (authenticated users only)

    A custom endpoint (`get-by-code/<code>`) allows retrieving a product
    directly by its unique code field.
    """

    queryset = Product.objects.all()
    serializer_class = ProductSerializer
    pagination_class = ProductPagination
    authentication_classes = [SessionAuthentication, TokenAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    @action(detail=False, methods=["get"], url_path=r"get-by-code/(?P<code>[\w-]+)")
    def get_by_code(self, request, code=None):
        """
        Retrieves a Product instance by its `code` value.
        """
        if not code:
            return Response(
                {"error": "Código inválido"}, status=status.HTTP_400_BAD_REQUEST
            )

        found = Product.objects.filter(code=code).first()

        if not found:
            return Response(
                {"error": "El producto con este código no existe"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        serializer = self.get_serializer(found)
        return Response(serializer.data)

    @action(detail=False, methods=["get"], url_path="low-stock/(?P<limit>\d+)")
    def low_stock(self, request, limit=None):
        """
        Return up to 150 products with stock less than or equal to `limit`, sorted by stock ascending
        """
        try:
            limit = float(limit)
        except ValueError:
            return Response(
                {"error": "Limite invalido"}, status=status.HTTP_400_BAD_REQUEST
            )

        products = Product.objects.filter(stock__lte=limit).order_by("stock")[:150]

        serializer = self.get_serializer(products, many=True)
        return Response(serializer.data)

    @action(
        detail=False, methods=["delete"], url_path="delete-by-code/(?P<code>[^/.]+)"
    )
    def destroy_by_code(self, request, code=None):
        """
        Deletes a product from the DB only if there's no sales associated with the product
        """
        product = Product.objects.filter(code=code).first()
        if not product:
            return Response(
                {
                    "success": False,
                    "error": f'No se encontró el producto con código "{code}".',
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        if SaleItem.objects.filter(product=product).exists():
            return Response(
                {
                    "success": False,
                    "error": "El producto tiene ventas asignadas, no puede ser eliminado.",
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        product.delete()
        return Response({"success": True}, status=status.HTTP_200_OK)

    @action(detail=False, methods=["patch"], url_path="patch-by-code/(?P<code>[^/.]+)")
    def patch_by_code(self, request, code=None):
        """
        Modifies a certain product
        """
        validate_response = ProductValidator.validate_data(request)
        if validate_response["success"] is False:
            return Response(
                {"success": False, "error": validate_response["error"] or ""},
                status=status.HTTP_400_BAD_REQUEST,
            )

        product = Product.objects.filter(code=code).first()

        if not product:
            return Response(
                {"success": False, "error": "No se encontro el producto a modificar"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        serializer = ProductSerializer(product, data=request.data, partial=True)

        if serializer.is_valid():
            for attr, value in serializer.validated_data.items():
                setattr(product, attr, value)

            product.last_modification = timezone.now()
            product.save(user=request.user)

            return Response({"success": True, "error": ""}, status=status.HTTP_200_OK)

        else:
            print(serializer.errors)
            return Response(
                {"success": False, "error": "El codigo ya existe para otro producto"},
                status=status.HTTP_400_BAD_REQUEST,
            )

    @action(detail=False, methods=["patch"], url_path="patch-selected-prices")
    def patch_selected(self, request):
        """
        Updates a list of selected products by code,
        with optional control for whether to include discounted products (in an active offer).
        """
        percentage = request.data.get("percentage")
        include_discounted = request.data.get("includeDiscounted", False)
        codes = request.data.get("codes", [])

        if not isinstance(codes, list) or not codes:
            return Response(
                {
                    "success": False,
                    "error": "Debe proporcionar una lista de códigos de productos",
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        updated_products = []

        for code in codes:
            product = Product.objects.filter(code=code).first()

            if not product:
                return Response(
                    {
                        "success": False,
                        "error": f"No se encontró el producto con código {code}",
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            if product.has_discount() and not include_discounted:
                continue

            if isinstance(percentage, (int, float)):
                product.sell_price = product.sell_price * (1 + (percentage / 100))
                product.last_modification = timezone.now()
                product.save(user=request.user)
                updated_products.append(product.code)

        return Response(
            {
                "success": True,
                "updated": updated_products,
            },
            status=status.HTTP_200_OK,
        )

    @action(detail=False, methods=["patch"], url_path="patch-all-prices")
    def patch_all(self, request):
        """
        Applies a percentage increase to all product prices,
        optionally including products that already have a discount.
        """
        percentage = request.data.get("percentage")
        include_discounted = request.data.get("includeDiscounted", False)

        products = Product.objects.all()

        for product in products:
            if product.has_discount() and not include_discounted:
                continue

            if isinstance(percentage, (int, float)):
                product.sell_price = product.sell_price * (1 + (percentage / 100))
                product.last_modification = timezone.now()
                product.save(user=request.user)

        return Response({"success": True}, status=status.HTTP_200_OK)

    @action(detail=False, methods=["get"], url_path="downloadExcel")
    def generate_excel(self, request):
        """
        Generates and returns an Excel file containing the full inventory,
        formatted for readability and including product and provider details.
        """
        try:
            base_dir = os.path.dirname(
                os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            )
            forms_dir = os.path.join(base_dir, "formularios")
            os.makedirs(forms_dir, exist_ok=True)
            file_name = "inventario.xlsx"
            file_path = os.path.join(forms_dir, file_name)

            data = []
            max_len = {
                "Código": len("Código"),
                "Nombre": len("Nombre"),
                "Stock": len("Stock"),
                "Precio de venta": len("Precio de venta"),
                "Precio de compra": len("Precio de compra"),
                "Proveedor": len("Proveedor"),
            }
            for product in Product.objects.all():
                provider_name = "No Registrado"
                if product.provider:
                    provider_name = product.provider.name
                item = {
                    "Código": product.code,
                    "Nombre": product.name,
                    "Stock": product.stock,
                    "Precio de venta": product.sell_price,
                    "Precio de compra": product.buy_price,
                    "Proveedor": provider_name,
                }
                data.append(item)
                for key, value in item.items():
                    if len(str(value)) > max_len[key]:
                        max_len[key] = len(str(value))

            wb = Workbook()
            ws = wb.active
            ws.title = "Inventario"

            headers = list(data[0].keys()) if data else []
            ws.append(headers)

            for i, header in enumerate(headers, start=1):
                col_letter = get_column_letter(i)
                ws.column_dimensions[col_letter].width = max_len[header] + 2

            fill = PatternFill(
                start_color="f1f1f1", end_color="f1f1f1", fill_type="solid"
            )

            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(bold=True)

            for idx, item in enumerate(data, start=1):
                row = [item[h] for h in headers]
                ws.append(row)
                if (idx + 1) % 2 == 0:
                    for col in range(1, len(headers) + 1):
                        cell = ws.cell(row=idx + 1, column=col)
                        cell.fill = fill
                        cell.alignment = Alignment(
                            horizontal="center", vertical="center"
                        )
                else:
                    for col in range(1, len(headers) + 1):
                        cell = ws.cell(row=idx + 1, column=col)
                        cell.alignment = Alignment(
                            horizontal="center", vertical="center"
                        )

            wb.save(file_path)

            return FileResponse(
                open(file_path, "rb"),
                as_attachment=True,
                filename=file_name,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        except Exception as e:
            return Response({"error": str(e)}, status=500)


class OfferViewSet(viewsets.ModelViewSet):
    """
    Provides CRUD operations for managing product offers.
    """

    queryset = Offer.objects.all()
    serializer_class = OfferSerializer
    pagination_class = OfferPagination

    def create(self, request, *args, **kwargs):
        """
        Creates a new offer after validating required fields and product IDs.
        """
        try:
            name = request.data.get("name")
            percentage = request.data.get("percentage")
            end_date = request.data.get("end_date")
            products_ids = request.data.get("products")

            if not name:
                return Response(
                    {"success": False, "error": "El nombre es obligatorio."}, 400
                )

            if Offer.objects.filter(name=name).exists():
                return Response(
                    {"success": False, "error": "Ya existe una oferta con ese nombre."},
                    400,
                )

            if percentage is None:
                return Response(
                    {"success": False, "error": "El porcentaje es obligatorio."}, 400
                )

            if end_date is None:
                return Response(
                    {
                        "success": False,
                        "error": "La fecha de finalización es obligatoria.",
                    },
                    400,
                )

            if not products_ids or not isinstance(products_ids, list):
                return Response(
                    {
                        "success": False,
                        "error": "Debe proporcionar una lista de productos.",
                    },
                    400,
                )

            products = Product.objects.filter(pk__in=products_ids)

            if products.count() != len(products_ids):
                return Response(
                    {"success": False, "error": "Uno o más productos no existen."}, 400
                )

            for product in products:
                if product.has_discount():
                    return Response(
                        {
                            "success": False,
                            "error": f"{product.name} ya tiene una oferta asignada",
                        },
                        400,
                    )

            offer = Offer.objects.create(
                name=name,
                percentage=percentage,
                end_date=end_date,
            )
            offer.products.set(products)

            return Response({"success": True, "error": ""}, 201)

        except Exception as e:
            return Response({"success": False, "error": str(e)}, 500)

    def destroy(self, request, *args, **kwargs):
        """
        Deletes an existing offer if it exists.
        """
        try:
            offer = self.get_object()
            offer.delete()
            return Response({"success": True, "error": ""}, 200)
        except Offer.DoesNotExist:
            return Response({"success": False, "error": "La oferta no existe."}, 404)
        except Exception as e:
            return Response({"success": False, "error": str(e)}, 500)

    def update(self, request, *args, **kwargs):
        """
        Updates an existing offer, including assigned products if provided.
        """
        try:
            offer = self.get_object()

            name = request.data.get("name", offer.name)
            percentage = request.data.get("percentage", offer.percentage)
            end_date = request.data.get("end_date", offer.end_date)
            products_ids = request.data.get("products", None)

            if Offer.objects.filter(name=name).exclude(id=offer.id).exists():
                return Response(
                    {
                        "success": False,
                        "error": "Ya existe otra oferta con ese nombre.",
                    },
                    400,
                )

            if products_ids is not None:
                if not isinstance(products_ids, list):
                    return Response(
                        {"success": False, "error": "Products debe ser una lista."}, 400
                    )

                products = Product.objects.filter(pk__in=products_ids)
                if products.count() != len(products_ids):
                    return Response(
                        {"success": False, "error": "Uno o más productos no existen."},
                        400,
                    )

                offer.products.set(products)

            offer.name = name
            offer.percentage = percentage
            offer.end_date = end_date
            offer.save()

            return Response({"success": True, "error": ""}, 200)

        except Offer.DoesNotExist:
            return Response({"success": False, "error": "La oferta no existe."}, 404)
        except Exception as e:
            return Response({"success": False, "error": str(e)}, 500)

    def list(self, request, *args, **kwargs):
        """
        Retrieves all offers with optional pagination support.
        """
        try:
            queryset = self.get_queryset()

            page = self.paginate_queryset(queryset)
            if page is not None:
                serializer = self.get_serializer(page, many=True)
                return self.get_paginated_response(
                    {"success": True, "error": "", "offers": serializer.data}
                )

            serializer = self.get_serializer(queryset, many=True)
            return Response({"success": True, "error": "", "offers": serializer.data})

        except Exception as e:
            return Response({"success": False, "error": str(e)}, 500)


class ProductSearchView(APIView):
    """
    Performs a relevance-based search across products by name, code, or stock.
    """

    def get(self, request):
        query = request.GET.get("q", "").strip().lower()
        if not query:
            return Response([], status=status.HTTP_200_OK)

        results = []
        for product in Product.objects.all():
            score = 0

            name = product.name.lower() if product.name else ""
            code = product.code.lower() if product.code else ""
            stock_str = str(product.stock) if product.stock is not None else ""
            category = product.category.name if product.category is not None else ""

            if query in name:
                score += 5
                if name.startswith(query):
                    score += 3

            if query in code:
                score += 4
                if code.startswith(query):
                    score += 2

            if query in category:
                score += 5
                if category.startswith(query):
                    score += 3

            if query == stock_str:
                score += 2

            if score > 0:
                results.append((score, product))

        results.sort(key=lambda tup: tup[0], reverse=True)
        matched_products = [p for _, p in results]

        serializer = ProductSerializer(matched_products, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
