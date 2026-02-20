from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import os
from django.conf import settings
from datetime import datetime

def get_nested_attr(obj, attr_path):
    """
    Helper to retrieve nested attributes using dunder notation (e.g., 'created_by__username').
    """
    try:
        for attr in attr_path.split("__"):
            obj = getattr(obj, attr)
            if obj is None:
                return ""
        return obj
    except AttributeError:
        return ""

def export_to_excel(filename: str, columns: list, queryset):
    """
    Generates an Excel file from a queryset with styling.
    
    filename: name of the file (without .xlsx)
    columns: list of tuples -> [("Header Title", "model_field__lookup"), ...]
    queryset: queryset or iterable
    
    Returns: (success: bool, file_path: str)
    """
    output_folder = os.path.join(settings.MEDIA_ROOT, "planillas_excel")
    os.makedirs(output_folder, exist_ok=True)
    
    file_name = f'{filename}.xlsx'
    file_path = os.path.join(output_folder, file_name)

    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    
    # Headers
    headers = [col[0] for col in columns]
    ws.append(headers)
    
    max_len = {header: len(str(header)) for header in headers}
    
    # Process data
    formatted_rows = []
    for obj in queryset:
        row = []
        for i, (_, field) in enumerate(columns):
            value = get_nested_attr(obj, field)
            
            if isinstance(value, datetime):
                value = value.strftime("%Y-%m-%d %H:%M:%S")
                
            str_value = str(value) if value is not None else ""
            row.append(str_value)
            
            # Update max length for column width
            header_key = headers[i]
            max_len[header_key] = max(max_len[header_key], len(str_value))
            
        formatted_rows.append(row)

    for i, header in enumerate(headers, start=1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = max_len[header] + 2

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True)

    fill = PatternFill(start_color="f1f1f1", end_color="f1f1f1", fill_type="solid")
    
    for idx, row_data in enumerate(formatted_rows, start=1):
        ws.append(row_data)
        
        current_row = idx + 1
        is_even = current_row % 2 == 0
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=current_row, column=col)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if is_even:
                cell.fill = fill

    try:
        wb.save(file_path)
        return True, file_path
    except Exception as e:
        print(f"Error saving excel: {e}")
        return False, 'Error al guardar el archivo. Verifique permisos o cierre el archivo si está abierto.'
